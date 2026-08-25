"""
월결산서 생성기 — 기존에 쓰던 "손익자료분석" 엑셀 구조를 그대로 재사용.

이 파일(monthly_closing_template.xlsx)에는 이미:
- 백데이터 시트 4개: 재무상태표 / 손익계산서 / 기간별손익계산서 / 기간별손익계산서(전년동기)
- 보고서 시트 5개: 인덱스 / 요약 / 재무상태표(내역) / 매출액대비 비율분석 / 전월대비증감
이렇게 구성되어 있고, 보고서 시트들은 백데이터 시트를 INDEX/MATCH 수식으로 참조한다.

매달 할 일은: 백데이터 시트 4개의 "내용"만 새로 업로드된 파일로 교체하고,
각 행마다 있는 매칭용 키 수식(과목명 정제 + 중복순번)을 새 행 개수에 맞게 다시 세팅하는 것.
보고서 시트는 손대지 않으므로 수기로 입력해둔 "내역"(비고) 칸도 그대로 보존된다.

업로드되는 백데이터는 최신 엑셀(.xlsx)뿐 아니라 회계프로그램에서 바로 내보낸
예전 방식(.xls)도 그대로 받을 수 있어야 하고, 계정과목 유무에 따라 행/열 개수가
매달 달라질 수 있으므로(예: 상반기만 있으면 8열, 연말이면 14열) 항상 "지금 실제
채워진 만큼만" 읽어서 처리한다.
"""
from __future__ import annotations
from io import BytesIO
import datetime
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

# 백데이터 시트별 설정: 어느 열까지가 실제 값이고, 이름/키 수식이 어느 열에 들어가는지, 몇 행부터 시작하는지
RAW_SHEET_CONFIG = {
    '재무상태표': {'data_start_col': 'A', 'data_end_col': 'E', 'name_col': 'F', 'key_col': 'G', 'start_row': 3},
    '손익계산서': {'data_start_col': 'A', 'data_end_col': 'E', 'name_col': 'F', 'key_col': 'G', 'start_row': 3},
    '기간별손익계산서': {'data_start_col': 'A', 'data_end_col': 'N', 'name_col': 'P', 'key_col': 'Q', 'start_row': 2},
    '기간별손익계산서(전년동기)': {'data_start_col': 'A', 'data_end_col': 'N', 'name_col': 'O', 'key_col': 'P', 'start_row': 2},
}


def _load_rows(file_obj) -> list[list]:
    """업로드된 백데이터 파일을 읽어서 2차원 리스트(행렬)로 반환.
    최신 엑셀(.xlsx, PK로 시작하는 zip 포맷)과 예전 엑셀(.xls, OLE2 포맷) 둘 다 지원.
    각 행의 길이는 원본 그대로(짧을 수 있음) — 컬럼이 모자란 부분은 나중에 None 처리."""
    if hasattr(file_obj, "read"):
        data = file_obj.read()
        try:
            file_obj.seek(0)
        except Exception:
            pass
    else:
        with open(file_obj, "rb") as f:
            data = f.read()

    if data[:2] == b"PK":
        # 최신 엑셀(.xlsx)
        wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
        ws = wb.worksheets[0]
        rows = []
        for row in ws.iter_rows():
            rows.append([c.value for c in row])
        return rows

    # 예전 엑셀(.xls) — 회계프로그램(더존 등)에서 바로 내보낸 형식
    import xlrd
    wb = xlrd.open_workbook(file_contents=data)
    ws = wb.sheet_by_index(0)
    rows = []
    for r in range(ws.nrows):
        row = []
        for c in range(ws.ncols):
            v = ws.cell_value(r, c)
            row.append(None if v == "" else v)
        rows.append(row)
    return rows


def _cell(rows: list[list], row_1idx: int, col_1idx: int):
    """1-indexed 행/열로 rows(0-indexed 행렬)에서 값 가져오기. 범위를 벗어나면 None."""
    r = row_1idx - 1
    c = col_1idx - 1
    if r < 0 or r >= len(rows):
        return None
    line = rows[r]
    if c < 0 or c >= len(line):
        return None
    return line[c]


def _clear_and_fill(dst_ws, cfg, src_rows: list[list]) -> dict:
    """raw 시트를 새 데이터로 채우고, {정규화된 계정명: 실제로 들어간 행 번호} 맵을 반환.
    (요약/재무상태표(내역) 시트 안에 일부 수식이 "몇 번째 행"을 통째로 하드코딩해서
    참조하는 곳이 있어서, 매달 계정 순서가 달라지면 엉뚱한 값을 가리키게 됨 —
    그 수식들을 다시 정확한 행 번호로 고쳐 쓰기 위해 이 맵이 필요함.)"""
    start_row = cfg['start_row']
    start_col = column_index_from_string(cfg['data_start_col'])
    end_col = column_index_from_string(cfg['data_end_col'])
    name_col = column_index_from_string(cfg['name_col'])
    key_col = column_index_from_string(cfg['key_col'])
    name_letter = cfg['name_col']
    a_letter = cfg['data_start_col']

    # 1) 기존 데이터/이름/키 영역 클리어 (넉넉하게 300행까지)
    clear_until = max(dst_ws.max_row, 300)
    for r in range(start_row, clear_until + 1):
        for c in range(start_col, end_col + 1):
            dst_ws.cell(row=r, column=c).value = None
        dst_ws.cell(row=r, column=name_col).value = None
        dst_ws.cell(row=r, column=key_col).value = None

    # 2) 새 파일의 값을 그대로 복사 + 이름/키 수식 재생성
    # (실제 업로드 파일의 행/열 개수는 계정과목 유무에 따라 매달 달라질 수 있어
    #  항상 그 파일에 실제로 들어있는 만큼만 읽는다 — _cell()이 범위 밖이면 None을 반환)
    row_map = {}
    src_max_row = len(src_rows)
    out_row = start_row
    for r in range(start_row, src_max_row + 1):
        row_has_value = any(
            _cell(src_rows, r, c) not in (None, '')
            for c in range(start_col, end_col + 1)
        )
        if not row_has_value:
            out_row += 1
            continue
        for c in range(start_col, end_col + 1):
            dst_ws.cell(row=out_row, column=c).value = _cell(src_rows, r, c)
        dst_ws.cell(row=out_row, column=name_col).value = (
            f'=IF({a_letter}{out_row}="","",SUBSTITUTE({a_letter}{out_row}," ",""))'
        )
        dst_ws.cell(row=out_row, column=key_col).value = (
            f'=IF({a_letter}{out_row}="","",{name_letter}{out_row}&"|"&COUNTIF(${name_letter}${start_row}:{name_letter}{out_row},{name_letter}{out_row}))'
        )
        label = _cell(src_rows, r, start_col)
        if label:
            row_map.setdefault(normalize_key(label), out_row)
        out_row += 1
    return row_map


# 요약/재무상태표(내역) 시트 안에서 "몇 번째 행"을 그대로 하드코딩해서 참조하는 셀들.
# (원본 파일에서는 우연히 그 행이 맞았지만, 백데이터마다 계정 순서가 달라질 수 있어서
#  매번 실제 위치를 다시 찾아서 정확한 행 번호로 고쳐 써야 함)
HARDCODED_ROW_REFS = [
    # (수식이 있는 시트, 셀범위, 참조 대상 원본시트, 찾을 계정명, 어떤 데이터소스에서 행번호를 구할지)
    {"target_sheet": "요약", "cells": [f"B{r}" for r in range(29, 41)],
     "source_sheet": "기간별손익계산서", "account": "Ⅰ.매출액", "kind": "index"},
    {"target_sheet": "요약", "cells": [f"C{r}" for r in range(29, 41)],
     "source_sheet": "기간별손익계산서", "account": "Ⅴ.영업이익", "kind": "index"},
    {"target_sheet": "재무상태표(내역)", "cells": ["A77"],
     "source_sheet": "손익계산서", "account": "Ⅹ.당기순이익", "kind": "direct", "direct_col": "C"},
    {"target_sheet": "재무상태표(내역)", "cells": ["A78"],
     "source_sheet": "손익계산서", "account": "Ⅹ.당기순이익", "kind": "direct", "direct_col": "E"},
]


def _fix_hardcoded_row_refs(wb, row_maps: dict) -> None:
    """HARDCODED_ROW_REFS에 정의된 셀들의 수식 안에 있는 "몇 번째 행" 숫자를,
    이번에 실제로 채워진 행 번호로 다시 써줌. row_maps: {원본시트명: {정규화계정명: 행번호}}"""
    import re
    for spec in HARDCODED_ROW_REFS:
        if spec["target_sheet"] not in wb.sheetnames:
            continue
        source_map = row_maps.get(spec["source_sheet"])
        if not source_map:
            continue  # 그 원본 시트를 이번에 새로 안 올리셨으면(=예전 데이터 유지) 손대지 않음
        account_key = normalize_key(spec["account"])
        new_row = source_map.get(account_key)
        if not new_row:
            continue  # 이번 달 데이터에 그 계정 자체가 없으면 건드리지 않음(원래 값 유지)

        ws = wb[spec["target_sheet"]]
        for cell_addr in spec["cells"]:
            cell = ws[cell_addr]
            formula = cell.value
            if not isinstance(formula, str) or not formula.startswith("="):
                continue
            if spec["kind"] == "index":
                # INDEX(시트!$C:$N, 33, N) 형태에서 "33"만 새 행 번호로 교체
                pattern = re.compile(
                    r'(INDEX\(' + re.escape(spec["source_sheet"]) + r'!\$?[A-Z]+:\$?[A-Z]+,\s*)\d+(\s*,)'
                )
                new_formula = pattern.sub(rf'\g<1>{new_row}\g<2>', formula)
            else:  # direct: 시트!C55 형태에서 행번호만 교체
                col = spec["direct_col"]
                pattern = re.compile(re.escape(spec["source_sheet"]) + r'!' + re.escape(col) + r'\d+')
                new_formula = pattern.sub(f'{spec["source_sheet"]}!{col}{new_row}', formula)
            cell.value = new_formula


def build_monthly_closing(
    template_path: str,
    uploaded_files: dict,
    base_date: datetime.date,
    prepared_date: datetime.date | None = None,
    remarks_override: dict | None = None,
    special_notes: dict | None = None,
) -> bytes:
    """
    uploaded_files: {"재무상태표": <파일경로 or BytesIO>, "손익계산서": ..., "기간별손익계산서": ..., "기간별손익계산서(전년동기)": ...}
    없는 키는 건너뛰고 기존 데이터를 그대로 둔다. 파일은 .xlsx/.xls 둘 다 가능.

    base_date: 결산기준일 — 이 월의 실적을 다루는지(예: 2026-07-31)
    prepared_date: 작성일자 — 실제로 이 보고서를 작성/보고한 날짜(예: 2026-09-05, 기준일자와 다를 수 있음)
    remarks_override: {account_key: note} — 재무상태표(내역) D열(내역/비고)에 덮어쓸 값들
    special_notes: {note_key: text} — 요약 시트 "특이사항" 등 고정 위치 수기 입력란에 덮어쓸 값들
    """
    wb = openpyxl.load_workbook(template_path)

    row_maps = {}
    for sheet_name, cfg in RAW_SHEET_CONFIG.items():
        if sheet_name not in uploaded_files or uploaded_files[sheet_name] is None:
            continue
        src_rows = _load_rows(uploaded_files[sheet_name])
        dst_ws = wb[sheet_name]
        row_maps[sheet_name] = _clear_and_fill(dst_ws, cfg, src_rows)

    # 요약/재무상태표(내역) 안에 "몇 번째 행"을 그대로 박아둔 수식들을,
    # 이번에 실제로 채워진 행 번호로 다시 맞춰줌 (계정 순서가 달라져도 안전하게)
    _fix_hardcoded_row_refs(wb, row_maps)

    # 기준일 갱신 — 이 값 하나로 인덱스/요약/재무상태표(내역)/매출액대비 비율분석/전월대비증감의
    # 모든 기간 문구가 자동으로 바뀜
    wb['인덱스']['K1'] = base_date
    # 작성일자 — 실제 보고서를 작성/보고한 날짜(기준일자와 다를 수 있음). 안 주면 오늘 날짜로.
    wb['인덱스']['K2'] = prepared_date or datetime.date.today()

    if remarks_override:
        apply_all_remarks(wb, remarks_override)
    if special_notes:
        apply_special_notes(wb, special_notes)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


# 계정과목과 무관하게, 고정된 위치에 있는 수기 입력란(예: 요약 시트의 "특이사항")
SPECIAL_NOTE_CELLS = {
    "summary_special_note": {"sheet": "요약", "cell": "A69"},
}


def apply_special_notes(wb, notes: dict) -> None:
    """notes: {note_key: text} — SPECIAL_NOTE_CELLS에 정의된 고정 셀에 그대로 덮어씀.
    여러 줄(줄바꿈)을 그대로 셀에 넣으면 엑셀에서도 줄바꿈되어 보임(wrap_text 설정되어 있음)."""
    for key, cfg in SPECIAL_NOTE_CELLS.items():
        if key not in notes:
            continue
        if cfg["sheet"] not in wb.sheetnames:
            continue
        wb[cfg["sheet"]][cfg["cell"]] = notes[key] or ""


def normalize_key(name) -> str:
    if name is None:
        return ""
    return "".join(str(name).split())


# 수기로 "내역"(비고)을 적는 시트들 — 계정과목 열, 내역 열, 실제 데이터가 있는 행 범위
REMARKS_SHEETS = {
    '재무상태표(내역)': {'account_col': 1, 'note_col': 4, 'row_range': (9, 80)},
    '전월대비증감': {'account_col': 1, 'note_col': 13, 'row_range': (7, 54)},
}


def extract_all_remarks(wb_or_path) -> dict:
    """내역(비고)이 있는 시트들에서 계정과목별 현재 값을 그대로 뽑아옴.
    같은 계정명(예: "대손충당금")이 한 시트 안에 여러 번 나올 수 있어서(매출채권 대손충당금,
    미수금 대손충당금 등), 이름만으로는 서로 구분이 안 됨 — 그래서 "이름|몇번째로 등장했는지"를
    키로 씀. 이건 원본 엑셀 수식이 이미 쓰고 있는 방식과 동일하고(COUNTIF 기반), 이 보고서
    시트 자체의 줄 순서는 매달 안 바뀌므로(백데이터만 바뀜) 매달 같은 항목을 정확히 가리킴.
    반환: {시트명: [{row, account_key, account_label, note}, ...]}"""
    wb = wb_or_path if hasattr(wb_or_path, "sheetnames") else openpyxl.load_workbook(wb_or_path, data_only=True)
    result = {}
    for sheet_name, cfg in REMARKS_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = []
        occurrence_count = {}
        for r in range(cfg['row_range'][0], cfg['row_range'][1] + 1):
            label = ws.cell(row=r, column=cfg['account_col']).value
            if not label or not str(label).strip() or str(label).strip().startswith('='):
                continue
            base_key = normalize_key(label)
            occurrence_count[base_key] = occurrence_count.get(base_key, 0) + 1
            note = ws.cell(row=r, column=cfg['note_col']).value
            rows.append({
                "row": r,
                "account_key": f"{base_key}|{occurrence_count[base_key]}",
                "account_label": str(label).strip(),
                "note": note if isinstance(note, str) else (str(note) if note is not None else None),
            })
        result[sheet_name] = rows
    return result


def apply_all_remarks(wb, remarks_map: dict) -> None:
    """remarks_map: {account_key(이름|몇번째): note} — 내역이 있는 모든 시트에 덮어씀."""
    for sheet_name, cfg in REMARKS_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        occurrence_count = {}
        for r in range(cfg['row_range'][0], cfg['row_range'][1] + 1):
            label = ws.cell(row=r, column=cfg['account_col']).value
            if not label or not str(label).strip() or str(label).strip().startswith('='):
                continue
            base_key = normalize_key(label)
            occurrence_count[base_key] = occurrence_count.get(base_key, 0) + 1
            key = f"{base_key}|{occurrence_count[base_key]}"
            if key in remarks_map:
                ws.cell(row=r, column=cfg['note_col']).value = remarks_map[key] or None


# 미리보기/보고서용 핵심지표 — 라이브러리(LibreOffice) 없이도 바로 계산해서 보여주기 위해
# 백데이터 파일을 직접 읽어서 계정명으로 찾음 (엑셀 수식과 동일한 매칭 방식)
_IS_TARGETS = {
    "매출액": "Ⅰ.매출액",
    "매출원가": "Ⅱ.매출원가",
    "매출총이익": "Ⅲ.매출총이익",
    "판매관리비": "Ⅳ.판매관리비",
    "영업이익": "Ⅴ.영업이익",
    "영업외수익": "Ⅵ.영업외수익",
    "영업외비용": "Ⅶ.영업외비용",
    "법인세차감전순이익": "Ⅷ.법인세비용차감전순이익",
    "당기순이익": "Ⅹ.당기순이익",
}
_BS_TARGETS = {
    "자산총계": "자산총계",
    "부채총계": "부채총계",
    "자본총계": "자본총계",
}


def _to_number(v):
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, str):
        s = v.strip().replace(",", "")
        if s in ("", "-"):
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _first_nonzero(*values):
    nums = [_to_number(v) for v in values]
    for n in nums:
        if n is not None and n != 0:
            return n
    for n in nums:
        if n is not None:
            return n
    return 0


def _scan_raw_sheet_for_targets(file_obj, targets: dict, value_cols=(2, 3)) -> dict:
    """file_obj: 업로드된 원본 파일(재무상태표 또는 손익계산서 형식, A=과목, B/C 또는 지정된 컬럼이 당기 금액).
    targets: {친숙한이름: 정규화된 계정명}"""
    src_rows = _load_rows(file_obj)
    key_to_friendly = {v: k for k, v in targets.items()}
    result = {}
    for row in src_rows:
        label = row[0] if row else None
        if not label:
            continue
        key = normalize_key(label)
        if key in key_to_friendly:
            b = row[value_cols[0] - 1] if len(row) >= value_cols[0] else None
            c = row[value_cols[1] - 1] if len(row) >= value_cols[1] else None
            result[key_to_friendly[key]] = _first_nonzero(b, c)
    return result


def _scan_col_b(file_obj, targets: dict) -> dict:
    """기간별손익계산서류 파일에서 B열(계, =채워진 월들의 합계)을 계정명으로 찾아서 가져옴."""
    src_rows = _load_rows(file_obj)
    key_to_friendly = {v: k for k, v in targets.items()}
    result = {}
    for row in src_rows:
        label = row[0] if row else None
        if not label:
            continue
        key = normalize_key(label)
        if key in key_to_friendly:
            b = row[1] if len(row) > 1 else None
            result[key_to_friendly[key]] = _to_number(b) or 0
    return result


def _monthly_trend(file_obj, label_targets=("매출액", "영업이익")) -> dict:
    """기간별손익계산서(당해)에서 월별(C~N열, 1~12월) 매출액/영업이익 추이를 뽑아옴.
    실제 채워진 달만큼만 값이 들어오고 나머지는 None(회계연도 진행 중이면 당연히 그럼)."""
    targets = {"매출액": "Ⅰ.매출액", "영업이익": "Ⅴ.영업이익"}
    src_rows = _load_rows(file_obj)
    key_to_friendly = {v: k for k, v in targets.items() if k in label_targets}
    result = {k: [None] * 12 for k in label_targets}
    for row in src_rows:
        label = row[0] if row else None
        if not label:
            continue
        key = normalize_key(label)
        if key in key_to_friendly:
            friendly = key_to_friendly[key]
            for m in range(12):
                col_idx = 2 + m  # 0-index: C열=index2 → 1월
                v = row[col_idx] if len(row) > col_idx else None
                result[friendly][m] = _to_number(v)
    return result


def compute_preview_summary(uploaded_files: dict) -> dict:
    """업로드된 손익계산서/재무상태표 원본 파일에서 핵심지표를 직접 계산 (재계산 없이 바로 미리보기용)."""
    summary = {}
    if uploaded_files.get("손익계산서"):
        try:
            summary.update(_scan_raw_sheet_for_targets(uploaded_files["손익계산서"], _IS_TARGETS, value_cols=(2, 3)))
        except Exception:
            pass
    if uploaded_files.get("재무상태표"):
        try:
            summary.update(_scan_raw_sheet_for_targets(uploaded_files["재무상태표"], _BS_TARGETS, value_cols=(2, 3)))
        except Exception:
            pass
    return summary


def compute_full_report_summary(uploaded_files: dict) -> dict:
    """인쇄용 보고서에 필요한 전체 요약(당기/전년동기/전기 손익 + 재무상태 + 월별추이)을 계산.
    엑셀 수식과 동일한 계정명 매칭 방식을 그대로 파이썬으로 재현 — LibreOffice 재계산 불필요."""
    result = {"income": {}, "balance": {}, "trend": {}}

    if uploaded_files.get("손익계산서"):
        try:
            cur = _scan_raw_sheet_for_targets(uploaded_files["손익계산서"], _IS_TARGETS, value_cols=(2, 3))
            prev_year = _scan_raw_sheet_for_targets(uploaded_files["손익계산서"], _IS_TARGETS, value_cols=(4, 5))
            for k in _IS_TARGETS:
                result["income"].setdefault(k, {})["당기"] = cur.get(k, 0)
                result["income"].setdefault(k, {})["전기"] = prev_year.get(k, 0)
        except Exception:
            pass

    if uploaded_files.get("기간별손익계산서(전년동기)"):
        try:
            same_period_last_year = _scan_col_b(uploaded_files["기간별손익계산서(전년동기)"], _IS_TARGETS)
            for k, v in same_period_last_year.items():
                result["income"].setdefault(k, {})["전년동기"] = v
        except Exception:
            pass

    if uploaded_files.get("재무상태표"):
        try:
            cur = _scan_raw_sheet_for_targets(uploaded_files["재무상태표"], _BS_TARGETS, value_cols=(2, 3))
            prev = _scan_raw_sheet_for_targets(uploaded_files["재무상태표"], _BS_TARGETS, value_cols=(4, 5))
            for k in _BS_TARGETS:
                result["balance"].setdefault(k, {})["당기"] = cur.get(k, 0)
                result["balance"].setdefault(k, {})["전기"] = prev.get(k, 0)
        except Exception:
            pass

    if uploaded_files.get("기간별손익계산서"):
        try:
            result["trend"] = _monthly_trend(uploaded_files["기간별손익계산서"])
        except Exception:
            pass

    return result
